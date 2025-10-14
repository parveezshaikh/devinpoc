#!/usr/bin/env python3
import argparse
import pandas as pd
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def read_source_data(source_path):
    df = pd.read_csv(source_path, encoding='utf-8-sig')
    return df


def read_spark_output(output_dir):
    part_files = glob.glob(os.path.join(output_dir, 'part-*.csv'))
    if not part_files:
        raise FileNotFoundError(f"No part files found in {output_dir}")
    
    dfs = []
    for file in part_files:
        df = pd.read_csv(file)
        dfs.append(df)
    
    combined_df = pd.concat(dfs, ignore_index=True)
    return combined_df


def test_case_1_row_count(source_df, output_df, filter_condition='Area < 300'):
    source_df['Area'] = pd.to_numeric(source_df['Area'], errors='coerce')
    
    filtered_source = source_df[source_df['Area'] < 300]
    
    source_count = len(filtered_source)
    output_count = len(output_df)
    
    passed = source_count == output_count
    
    result = {
        'test_name': 'Test Case 1: Row Count Match',
        'status': 'PASS' if passed else 'FAIL',
        'details': f'Source (filtered): {source_count} rows, Output: {output_count} rows',
        'source_count': source_count,
        'output_count': output_count
    }
    
    return result


def test_case_2_area_validation(output_df):
    output_df['Area'] = pd.to_numeric(output_df['Area'], errors='coerce')
    
    violations = output_df[output_df['Area'] >= 300]
    violation_count = len(violations)
    
    passed = violation_count == 0
    
    result = {
        'test_name': 'Test Case 2: Area < 300 Validation',
        'status': 'PASS' if passed else 'FAIL',
        'details': f'Violations: {violation_count} (Expected: 0)',
        'violation_count': violation_count
    }
    
    return result


def test_case_3_balance_sum(source_df, output_df):
    source_df['Area'] = pd.to_numeric(source_df['Area'], errors='coerce')
    source_df['CurrentAccountBalance'] = pd.to_numeric(source_df['CurrentAccountBalance'], errors='coerce')
    
    filtered_source = source_df[source_df['Area'] < 300]
    
    source_sum = filtered_source['CurrentAccountBalance'].fillna(0).sum()
    
    output_df['CurrentAccountBalance'] = pd.to_numeric(output_df['CurrentAccountBalance'], errors='coerce')
    output_sum = output_df['CurrentAccountBalance'].fillna(0).sum()
    
    tolerance = 0.01
    passed = abs(source_sum - output_sum) < tolerance
    
    result = {
        'test_name': 'Test Case 3: CurrentAccountBalance Sum Match',
        'status': 'PASS' if passed else 'FAIL',
        'details': f'Source (filtered): {source_sum:.2f}, Output: {output_sum:.2f}',
        'source_sum': source_sum,
        'output_sum': output_sum
    }
    
    return result


def test_case_4_schema_validation(output_df):
    expected_columns = ['Country', 'Area', 'BirthRate', 'CurrentAccountBalance', 
                       'DeathRate', 'DebtExternal', 'ElectricityConsumption', 
                       'ElectricityProduction', 'Exports', 'GDP']
    
    missing_columns = [col for col in expected_columns if col not in output_df.columns]
    extra_columns = [col for col in output_df.columns if col not in expected_columns]
    
    passed = len(missing_columns) == 0 and len(extra_columns) == 0
    
    if passed:
        details = 'All expected columns present, no extra columns'
    else:
        details = f'Missing: {missing_columns}, Extra: {extra_columns}'
    
    result = {
        'test_name': 'Test Case 4: Schema Validation',
        'status': 'PASS' if passed else 'FAIL',
        'details': details,
        'missing_columns': missing_columns,
        'extra_columns': extra_columns
    }
    
    return result


def test_case_5_data_type_validation(output_df):
    numeric_columns = ['Area', 'BirthRate', 'CurrentAccountBalance', 'DeathRate', 
                      'DebtExternal', 'ElectricityConsumption', 'ElectricityProduction', 
                      'Exports', 'GDP']
    
    validation_results = {}
    all_passed = True
    
    for col in numeric_columns:
        if col in output_df.columns:
            try:
                pd.to_numeric(output_df[col], errors='coerce')
                validation_results[col] = 'Valid'
            except Exception as e:
                validation_results[col] = f'Invalid: {str(e)}'
                all_passed = False
        else:
            validation_results[col] = 'Missing'
            all_passed = False
    
    result = {
        'test_name': 'Test Case 5: Data Type Validation',
        'status': 'PASS' if all_passed else 'FAIL',
        'details': f'Numeric columns validation: {len([v for v in validation_results.values() if v == "Valid"])}/{len(numeric_columns)} valid',
        'validation_results': validation_results
    }
    
    return result


def test_case_6_null_value_analysis(source_df, output_df):
    source_df['Area'] = pd.to_numeric(source_df['Area'], errors='coerce')
    filtered_source = source_df[source_df['Area'] < 300]
    
    null_comparison = {}
    all_match = True
    
    for col in filtered_source.columns:
        if col in output_df.columns:
            source_nulls = filtered_source[col].isna().sum()
            output_nulls = output_df[col].isna().sum()
            null_comparison[col] = {
                'source_nulls': source_nulls,
                'output_nulls': output_nulls,
                'match': source_nulls == output_nulls
            }
            if source_nulls != output_nulls:
                all_match = False
    
    result = {
        'test_name': 'Test Case 6: Null Value Analysis',
        'status': 'PASS' if all_match else 'FAIL',
        'details': f'Null count match for all columns: {all_match}',
        'null_comparison': null_comparison
    }
    
    return result


def test_case_7_boundary_upper(output_df):
    output_df['Area'] = pd.to_numeric(output_df['Area'], errors='coerce')
    
    boundary_violations = output_df[output_df['Area'] == 300]
    max_area = output_df['Area'].max()
    
    passed = len(boundary_violations) == 0 and max_area < 300
    
    result = {
        'test_name': 'Test Case 7: Boundary Testing - Upper Bound',
        'status': 'PASS' if passed else 'FAIL',
        'details': f'Max Area: {max_area:.2f}, Area=300 count: {len(boundary_violations)}',
        'max_area': max_area,
        'boundary_violations': len(boundary_violations)
    }
    
    return result


def test_case_8_boundary_lower(output_df):
    output_df['Area'] = pd.to_numeric(output_df['Area'], errors='coerce')
    
    min_area = output_df['Area'].min()
    zero_or_small = output_df[output_df['Area'] <= 1]
    
    passed = min_area >= 0
    
    result = {
        'test_name': 'Test Case 8: Boundary Testing - Lower Bound',
        'status': 'PASS' if passed else 'FAIL',
        'details': f'Min Area: {min_area:.2f}, Area<=1 count: {len(zero_or_small)}',
        'min_area': min_area,
        'small_values_count': len(zero_or_small)
    }
    
    return result


def test_case_9_duplicate_detection(output_df):
    if 'Country' not in output_df.columns:
        return {
            'test_name': 'Test Case 9: Duplicate Country Detection',
            'status': 'FAIL',
            'details': 'Country column missing',
            'duplicates': []
        }
    
    duplicates = output_df[output_df.duplicated(subset=['Country'], keep=False)]
    duplicate_countries = duplicates['Country'].unique().tolist()
    
    passed = len(duplicates) == 0
    
    result = {
        'test_name': 'Test Case 9: Duplicate Country Detection',
        'status': 'PASS' if passed else 'FAIL',
        'details': f'Duplicate countries: {len(duplicate_countries)}',
        'duplicates': duplicate_countries
    }
    
    return result


def test_case_10_filter_completeness(output_df):
    output_df['Area'] = pd.to_numeric(output_df['Area'], errors='coerce')
    
    valid_records = output_df[output_df['Area'] < 300]
    invalid_records = output_df[output_df['Area'] >= 300]
    
    total = len(output_df)
    valid_pct = (len(valid_records) / total * 100) if total > 0 else 0
    
    passed = len(invalid_records) == 0
    
    result = {
        'test_name': 'Test Case 10: Filter Completeness',
        'status': 'PASS' if passed else 'FAIL',
        'details': f'{valid_pct:.2f}% records have Area < 300, Invalid: {len(invalid_records)}',
        'valid_records': len(valid_records),
        'invalid_records': len(invalid_records)
    }
    
    return result


def test_case_11_statistical_area(source_df, output_df):
    source_df['Area'] = pd.to_numeric(source_df['Area'], errors='coerce')
    filtered_source = source_df[source_df['Area'] < 300]
    output_df['Area'] = pd.to_numeric(output_df['Area'], errors='coerce')
    
    source_stats = {
        'mean': filtered_source['Area'].mean(),
        'median': filtered_source['Area'].median(),
        'min': filtered_source['Area'].min(),
        'max': filtered_source['Area'].max()
    }
    
    output_stats = {
        'mean': output_df['Area'].mean(),
        'median': output_df['Area'].median(),
        'min': output_df['Area'].min(),
        'max': output_df['Area'].max()
    }
    
    tolerance = 0.01
    passed = all([
        abs(source_stats['mean'] - output_stats['mean']) < tolerance,
        abs(source_stats['median'] - output_stats['median']) < tolerance,
        abs(source_stats['min'] - output_stats['min']) < tolerance,
        abs(source_stats['max'] - output_stats['max']) < tolerance
    ])
    
    result = {
        'test_name': 'Test Case 11: Statistical Validation - Area Column',
        'status': 'PASS' if passed else 'FAIL',
        'details': f'Mean: {output_stats["mean"]:.2f}, Median: {output_stats["median"]:.2f}, Min: {output_stats["min"]:.2f}, Max: {output_stats["max"]:.2f}',
        'source_stats': source_stats,
        'output_stats': output_stats
    }
    
    return result


def test_case_12_statistical_gdp(source_df, output_df):
    source_df['Area'] = pd.to_numeric(source_df['Area'], errors='coerce')
    source_df['GDP'] = pd.to_numeric(source_df['GDP'], errors='coerce')
    filtered_source = source_df[source_df['Area'] < 300]
    output_df['GDP'] = pd.to_numeric(output_df['GDP'], errors='coerce')
    
    source_sum = filtered_source['GDP'].fillna(0).sum()
    source_mean = filtered_source['GDP'].mean()
    output_sum = output_df['GDP'].fillna(0).sum()
    output_mean = output_df['GDP'].mean()
    
    tolerance = 0.01
    passed = abs(source_sum - output_sum) < tolerance and abs(source_mean - output_mean) < tolerance
    
    result = {
        'test_name': 'Test Case 12: Statistical Validation - GDP Column',
        'status': 'PASS' if passed else 'FAIL',
        'details': f'Sum: {output_sum:.2f}, Mean: {output_mean:.2f}',
        'source_sum': source_sum,
        'source_mean': source_mean,
        'output_sum': output_sum,
        'output_mean': output_mean
    }
    
    return result


def test_case_13_data_consistency(source_df, output_df):
    source_df['Area'] = pd.to_numeric(source_df['Area'], errors='coerce')
    filtered_source = source_df[source_df['Area'] < 300]
    
    consistency_check = {}
    all_match = True
    
    for col in filtered_source.columns:
        if col in output_df.columns:
            source_count = filtered_source[col].notna().sum()
            output_count = output_df[col].notna().sum()
            match = source_count == output_count
            consistency_check[col] = {
                'source': source_count,
                'output': output_count,
                'match': match
            }
            if not match:
                all_match = False
    
    result = {
        'test_name': 'Test Case 13: Data Consistency - Non-Null Counts',
        'status': 'PASS' if all_match else 'FAIL',
        'details': f'Non-null counts match for all columns: {all_match}',
        'consistency_check': consistency_check
    }
    
    return result


def test_case_14_output_structure(output_dir):
    import os
    
    checks = {
        'directory_exists': os.path.exists(output_dir),
        'has_part_files': False,
        'has_success_file': False
    }
    
    if checks['directory_exists']:
        files = os.listdir(output_dir)
        checks['has_part_files'] = any(f.startswith('part-') and f.endswith('.csv') for f in files)
        checks['has_success_file'] = '_SUCCESS' in files
    
    passed = all(checks.values())
    
    result = {
        'test_name': 'Test Case 14: Output File Structure Validation',
        'status': 'PASS' if passed else 'FAIL',
        'details': f'Dir exists: {checks["directory_exists"]}, Part files: {checks["has_part_files"]}, Success file: {checks["has_success_file"]}',
        'checks': checks
    }
    
    return result


def test_case_15_row_integrity(source_df, output_df):
    source_df['Area'] = pd.to_numeric(source_df['Area'], errors='coerce')
    filtered_source = source_df[source_df['Area'] < 300]
    
    source_count = len(filtered_source)
    output_count = len(output_df)
    
    no_data_loss = output_count >= source_count
    no_extra_rows = output_count <= source_count
    perfect_match = output_count == source_count
    
    passed = perfect_match
    
    result = {
        'test_name': 'Test Case 15: Row Integrity Check',
        'status': 'PASS' if passed else 'FAIL',
        'details': f'No data loss: {no_data_loss}, No extra rows: {no_extra_rows}, Perfect match: {perfect_match}',
        'source_count': source_count,
        'output_count': output_count,
        'difference': output_count - source_count
    }
    
    return result


def test_case_16_numeric_range(output_df):
    numeric_columns = ['Area', 'BirthRate', 'DeathRate', 'DebtExternal', 
                      'ElectricityConsumption', 'ElectricityProduction', 
                      'Exports', 'GDP']
    
    range_checks = {}
    all_valid = True
    
    for col in numeric_columns:
        if col in output_df.columns:
            output_df[col] = pd.to_numeric(output_df[col], errors='coerce')
            col_data = output_df[col].dropna()
            
            if len(col_data) > 0:
                min_val = col_data.min()
                max_val = col_data.max()
                has_negative = (min_val < 0) if col != 'CurrentAccountBalance' else False
                
                range_checks[col] = {
                    'min': min_val,
                    'max': max_val,
                    'has_negative': has_negative,
                    'valid': not has_negative
                }
                
                if has_negative:
                    all_valid = False
    
    result = {
        'test_name': 'Test Case 16: Numeric Range Validation',
        'status': 'PASS' if all_valid else 'FAIL',
        'details': f'All numeric ranges valid: {all_valid}',
        'range_checks': range_checks
    }
    
    return result


def column_level_validation(source_df, output_df):
    source_df['Area'] = pd.to_numeric(source_df['Area'], errors='coerce')
    filtered_source = source_df[source_df['Area'] < 300]
    
    results = []
    
    for column in filtered_source.columns:
        row_data = {
            'Column': column,
            'Source_Count': filtered_source[column].notna().sum(),
            'Output_Count': output_df[column].notna().sum() if column in output_df.columns else 0,
            'Source_Sum': '',
            'Output_Sum': '',
            'Match_Status': ''
        }
        
        if column in output_df.columns:
            source_numeric = pd.to_numeric(filtered_source[column], errors='coerce')
            output_numeric = pd.to_numeric(output_df[column], errors='coerce')
            
            if source_numeric.notna().sum() > 0:
                source_sum = source_numeric.fillna(0).sum()
                output_sum = output_numeric.fillna(0).sum()
                
                row_data['Source_Sum'] = f'{source_sum:.2f}'
                row_data['Output_Sum'] = f'{output_sum:.2f}'
                
                count_match = row_data['Source_Count'] == row_data['Output_Count']
                sum_match = abs(source_sum - output_sum) < 0.01
                
                if count_match and sum_match:
                    row_data['Match_Status'] = 'PASS'
                elif count_match:
                    row_data['Match_Status'] = 'COUNT_PASS'
                else:
                    row_data['Match_Status'] = 'FAIL'
            else:
                count_match = row_data['Source_Count'] == row_data['Output_Count']
                row_data['Match_Status'] = 'PASS' if count_match else 'FAIL'
        else:
            row_data['Match_Status'] = 'MISSING_COLUMN'
        
        results.append(row_data)
    
    return pd.DataFrame(results)


def generate_statistical_analysis(source_df, output_df):
    source_df['Area'] = pd.to_numeric(source_df['Area'], errors='coerce')
    filtered_source = source_df[source_df['Area'] < 300]
    
    numeric_columns = ['Area', 'BirthRate', 'DeathRate', 'DebtExternal', 
                      'ElectricityConsumption', 'ElectricityProduction', 
                      'Exports', 'GDP']
    
    stats_results = []
    
    for col in numeric_columns:
        if col in filtered_source.columns and col in output_df.columns:
            source_col = pd.to_numeric(filtered_source[col], errors='coerce')
            output_col = pd.to_numeric(output_df[col], errors='coerce')
            
            row_data = {
                'Column': col,
                'Source_Mean': f'{source_col.mean():.2f}' if source_col.notna().sum() > 0 else 'N/A',
                'Output_Mean': f'{output_col.mean():.2f}' if output_col.notna().sum() > 0 else 'N/A',
                'Source_Median': f'{source_col.median():.2f}' if source_col.notna().sum() > 0 else 'N/A',
                'Output_Median': f'{output_col.median():.2f}' if output_col.notna().sum() > 0 else 'N/A',
                'Source_Min': f'{source_col.min():.2f}' if source_col.notna().sum() > 0 else 'N/A',
                'Output_Min': f'{output_col.min():.2f}' if output_col.notna().sum() > 0 else 'N/A',
                'Source_Max': f'{source_col.max():.2f}' if source_col.notna().sum() > 0 else 'N/A',
                'Output_Max': f'{output_col.max():.2f}' if output_col.notna().sum() > 0 else 'N/A'
            }
            
            stats_results.append(row_data)
    
    return pd.DataFrame(stats_results)


def generate_data_quality_analysis(output_df, null_result, duplicate_result, range_result):
    quality_results = []
    
    for col in output_df.columns:
        null_count = output_df[col].isna().sum()
        total_count = len(output_df)
        null_pct = (null_count / total_count * 100) if total_count > 0 else 0
        
        row_data = {
            'Column': col,
            'Total_Records': total_count,
            'Null_Count': null_count,
            'Null_Percentage': f'{null_pct:.2f}%',
            'Non_Null_Count': total_count - null_count
        }
        
        if col in ['Area', 'BirthRate', 'DeathRate', 'DebtExternal', 
                  'ElectricityConsumption', 'ElectricityProduction', 'Exports', 'GDP']:
            output_df[col] = pd.to_numeric(output_df[col], errors='coerce')
            numeric_data = output_df[col].dropna()
            if len(numeric_data) > 0:
                row_data['Min_Value'] = f'{numeric_data.min():.2f}'
                row_data['Max_Value'] = f'{numeric_data.max():.2f}'
            else:
                row_data['Min_Value'] = 'N/A'
                row_data['Max_Value'] = 'N/A'
        else:
            row_data['Min_Value'] = 'N/A'
            row_data['Max_Value'] = 'N/A'
        
        quality_results.append(row_data)
    
    df = pd.DataFrame(quality_results)
    
    summary_row = {
        'Column': 'SUMMARY',
        'Total_Records': f'Duplicates: {len(duplicate_result.get("duplicates", []))}',
        'Null_Count': '',
        'Null_Percentage': '',
        'Non_Null_Count': '',
        'Min_Value': '',
        'Max_Value': ''
    }
    
    df = pd.concat([df, pd.DataFrame([summary_row])], ignore_index=True)
    
    return df


def generate_excel_report(test_results, column_results, statistical_data, data_quality_data, output_path):
    wb = Workbook()
    
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    ws_summary['A1'] = 'Test Case'
    ws_summary['B1'] = 'Status'
    ws_summary['C1'] = 'Details'
    
    for cell in ['A1', 'B1', 'C1']:
        ws_summary[cell].font = Font(bold=True)
        ws_summary[cell].alignment = Alignment(horizontal='left')
    
    for idx, result in enumerate(test_results, start=2):
        ws_summary[f'A{idx}'] = result['test_name']
        ws_summary[f'B{idx}'] = result['status']
        ws_summary[f'C{idx}'] = result['details']
    
    ws_summary.column_dimensions['A'].width = 50
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 70
    
    ws_details = wb.create_sheet(title="Column Details")
    
    for r_idx, row in enumerate(dataframe_to_rows(column_results, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_details.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left')
    
    for column in ws_details.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_details.column_dimensions[column_letter].width = adjusted_width
    
    ws_stats = wb.create_sheet(title="Statistical Analysis")
    
    for r_idx, row in enumerate(dataframe_to_rows(statistical_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_stats.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left')
    
    for column in ws_stats.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_stats.column_dimensions[column_letter].width = adjusted_width
    
    ws_quality = wb.create_sheet(title="Data Quality")
    
    for r_idx, row in enumerate(dataframe_to_rows(data_quality_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_quality.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left')
    
    for column in ws_quality.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_quality.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    print(f"Excel report generated: {output_path}")


def main():
    parser = argparse.ArgumentParser(description='Test pipeline data transformation')
    parser.add_argument('--source', required=True, help='Path to source CSV file')
    parser.add_argument('--output', required=True, help='Path to output directory (Spark output)')
    
    args = parser.parse_args()
    
    print(f"Loading source data from: {args.source}")
    source_df = read_source_data(args.source)
    print(f"Source data loaded: {len(source_df)} rows")
    
    print(f"Loading output data from: {args.output}")
    output_df = read_spark_output(args.output)
    print(f"Output data loaded: {len(output_df)} rows")
    
    print("\n" + "="*70)
    print("RUNNING COMPREHENSIVE TEST SUITE (16 Test Cases)")
    print("="*70)
    
    print("\nRunning Test Case 1: Row Count Match...")
    tc1_result = test_case_1_row_count(source_df, output_df)
    print(f"  {tc1_result['status']}: {tc1_result['details']}")
    
    print("\nRunning Test Case 2: Area < 300 Validation...")
    tc2_result = test_case_2_area_validation(output_df)
    print(f"  {tc2_result['status']}: {tc2_result['details']}")
    
    print("\nRunning Test Case 3: CurrentAccountBalance Sum Match...")
    tc3_result = test_case_3_balance_sum(source_df, output_df)
    print(f"  {tc3_result['status']}: {tc3_result['details']}")
    
    print("\nRunning Test Case 4: Schema Validation...")
    tc4_result = test_case_4_schema_validation(output_df)
    print(f"  {tc4_result['status']}: {tc4_result['details']}")
    
    print("\nRunning Test Case 5: Data Type Validation...")
    tc5_result = test_case_5_data_type_validation(output_df)
    print(f"  {tc5_result['status']}: {tc5_result['details']}")
    
    print("\nRunning Test Case 6: Null Value Analysis...")
    tc6_result = test_case_6_null_value_analysis(source_df.copy(), output_df.copy())
    print(f"  {tc6_result['status']}: {tc6_result['details']}")
    
    print("\nRunning Test Case 7: Boundary Testing - Upper Bound...")
    tc7_result = test_case_7_boundary_upper(output_df.copy())
    print(f"  {tc7_result['status']}: {tc7_result['details']}")
    
    print("\nRunning Test Case 8: Boundary Testing - Lower Bound...")
    tc8_result = test_case_8_boundary_lower(output_df.copy())
    print(f"  {tc8_result['status']}: {tc8_result['details']}")
    
    print("\nRunning Test Case 9: Duplicate Country Detection...")
    tc9_result = test_case_9_duplicate_detection(output_df.copy())
    print(f"  {tc9_result['status']}: {tc9_result['details']}")
    
    print("\nRunning Test Case 10: Filter Completeness...")
    tc10_result = test_case_10_filter_completeness(output_df.copy())
    print(f"  {tc10_result['status']}: {tc10_result['details']}")
    
    print("\nRunning Test Case 11: Statistical Validation - Area Column...")
    tc11_result = test_case_11_statistical_area(source_df.copy(), output_df.copy())
    print(f"  {tc11_result['status']}: {tc11_result['details']}")
    
    print("\nRunning Test Case 12: Statistical Validation - GDP Column...")
    tc12_result = test_case_12_statistical_gdp(source_df.copy(), output_df.copy())
    print(f"  {tc12_result['status']}: {tc12_result['details']}")
    
    print("\nRunning Test Case 13: Data Consistency - Non-Null Counts...")
    tc13_result = test_case_13_data_consistency(source_df.copy(), output_df.copy())
    print(f"  {tc13_result['status']}: {tc13_result['details']}")
    
    print("\nRunning Test Case 14: Output File Structure Validation...")
    tc14_result = test_case_14_output_structure(args.output)
    print(f"  {tc14_result['status']}: {tc14_result['details']}")
    
    print("\nRunning Test Case 15: Row Integrity Check...")
    tc15_result = test_case_15_row_integrity(source_df.copy(), output_df.copy())
    print(f"  {tc15_result['status']}: {tc15_result['details']}")
    
    print("\nRunning Test Case 16: Numeric Range Validation...")
    tc16_result = test_case_16_numeric_range(output_df.copy())
    print(f"  {tc16_result['status']}: {tc16_result['details']}")
    
    print("\nGenerating column-level validation...")
    column_results = column_level_validation(source_df, output_df)
    
    print("\nGenerating statistical analysis data...")
    statistical_data = generate_statistical_analysis(source_df.copy(), output_df.copy())
    
    print("\nGenerating data quality analysis...")
    data_quality_data = generate_data_quality_analysis(output_df.copy(), tc6_result, tc9_result, tc16_result)
    
    test_results = [
        tc1_result, tc2_result, tc3_result, tc4_result, tc5_result,
        tc6_result, tc7_result, tc8_result, tc9_result, tc10_result,
        tc11_result, tc12_result, tc13_result, tc14_result, tc15_result, tc16_result
    ]
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    result_dir = os.path.join(script_dir, '..', 'result')
    os.makedirs(result_dir, exist_ok=True)
    
    output_excel = os.path.join(result_dir, 'result51.xlsx')
    
    print(f"\nGenerating Excel report: {output_excel}")
    generate_excel_report(test_results, column_results, statistical_data, data_quality_data, output_excel)
    
    print("\n" + "="*70)
    print("COMPREHENSIVE TEST SUMMARY")
    print("="*70)
    for result in test_results:
        status_symbol = "✓" if result['status'] == 'PASS' else "✗"
        print(f"{status_symbol} {result['test_name']}: {result['status']}")
    print("="*70)
    
    passed_count = sum(1 for r in test_results if r['status'] == 'PASS')
    total_count = len(test_results)
    
    print(f"\nTest Results: {passed_count}/{total_count} PASSED")
    
    all_passed = all(r['status'] == 'PASS' for r in test_results)
    if all_passed:
        print("\n✓ All tests PASSED!")
    else:
        print(f"\n✗ {total_count - passed_count} test(s) FAILED!")
    
    return 0 if all_passed else 1


if __name__ == '__main__':
    exit(main())
